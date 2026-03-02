import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class MaterialSpecifier:
    def __init__(self):
        self.output_dir = os.path.join("assets", "engineering")
        os.makedirs(self.output_dir, exist_ok=True)

    def generate_boq(self, project_name, items):
        """Generate a professional Bill of Quantities Excel workbook."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Bill of Quantities"

        # ---- STYLES ----
        header_fill = PatternFill(start_color="0D1117", end_color="0D1117", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        brand_font = Font(name="Segoe UI", size=18, bold=True, color="58A6FF")
        sub_font = Font(name="Segoe UI", size=10, color="8B949E")
        cat_fill = PatternFill(start_color="161B22", end_color="161B22", fill_type="solid")
        cat_font = Font(name="Segoe UI", size=10, bold=True, color="58A6FF")
        data_font = Font(name="Segoe UI", size=10, color="C9D1D9")
        money_font = Font(name="Segoe UI", size=10, color="3FB950")
        total_fill = PatternFill(start_color="1F6FEB", end_color="1F6FEB", fill_type="solid")
        total_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        alt_fill = PatternFill(start_color="0D1117", end_color="0D1117", fill_type="solid")
        even_fill = PatternFill(start_color="161B22", end_color="161B22", fill_type="solid")
        thin_border = Border(
            bottom=Side(style='thin', color='30363D')
        )
        center = Alignment(horizontal='center', vertical='center')
        left = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')

        # ---- HEADER SECTION ----
        ws.merge_cells('A1:J1')
        ws['A1'] = '⚡ VOLTS DESIGN'
        ws['A1'].font = brand_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = left

        ws.merge_cells('A2:J2')
        ws['A2'] = f'Bill of Quantities — {project_name}'
        ws['A2'].font = Font(name="Segoe UI", size=13, bold=True, color="C9D1D9")
        ws['A2'].fill = header_fill

        ws.merge_cells('A3:J3')
        ws['A3'] = f'Generated: {datetime.datetime.now().strftime("%B %d, %Y %I:%M %p")}  |  Rev 1.0'
        ws['A3'].font = sub_font
        ws['A3'].fill = header_fill

        # Empty row
        ws.merge_cells('A4:J4')
        ws['A4'].fill = header_fill

        # ---- COLUMN HEADERS ----
        headers = ['#', 'Description', 'Specification', 'Manufacturer', 'Unit', 'Qty', 'Rate ($)', 'Amount ($)', 'Room', 'Status']
        col_widths = [5, 30, 20, 18, 8, 7, 12, 14, 16, 10]
        
        for i, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=5, column=i, value=h)
            cell.font = header_font
            cell.fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
            cell.alignment = center
            ws.column_dimensions[get_column_letter(i)].width = w

        # ---- DATA ROWS ----
        # Group by category
        categories = {}
        for item in items:
            cat = item.get('category', 'General')
            if cat not in categories:
                categories[cat] = []
            categories[cat].append(item)

        row = 6
        item_num = 0
        subtotals = {}

        for cat_name, cat_items in categories.items():
            # Category header
            ws.merge_cells(f'A{row}:J{row}')
            ws.cell(row=row, column=1, value=f'▸ {cat_name.upper()}')
            ws.cell(row=row, column=1).font = cat_font
            ws.cell(row=row, column=1).fill = cat_fill
            ws.cell(row=row, column=1).alignment = left
            row += 1

            cat_total = 0
            for idx, item in enumerate(cat_items):
                item_num += 1
                amount = round(item.get('qty', 0) * item.get('rate', 0), 2)
                cat_total += amount
                fill = even_fill if idx % 2 == 0 else alt_fill

                values = [
                    item_num,
                    item.get('item', ''),
                    item.get('specification', item.get('material', '')),
                    item.get('manufacturer', ''),
                    item.get('unit', 'pcs'),
                    item.get('qty', 0),
                    item.get('rate', item.get('est_cost', 0)),
                    amount,
                    item.get('room', ''),
                    item.get('status', 'Pending')
                ]

                for col, val in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.font = money_font if col in [7, 8] else data_font
                    cell.fill = fill
                    cell.border = thin_border
                    if col in [1, 5, 6]:
                        cell.alignment = center
                    elif col in [7, 8]:
                        cell.alignment = right_align
                        cell.number_format = '#,##0.00'
                    else:
                        cell.alignment = left

                row += 1

            # Category subtotal
            ws.merge_cells(f'A{row}:G{row}')
            ws.cell(row=row, column=1, value=f'Subtotal — {cat_name}')
            ws.cell(row=row, column=1).font = Font(name="Segoe UI", size=10, bold=True, color="8B949E")
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='right', vertical='center')
            ws.cell(row=row, column=1).fill = cat_fill
            
            sub_cell = ws.cell(row=row, column=8, value=cat_total)
            sub_cell.font = Font(name="Segoe UI", size=10, bold=True, color="3FB950")
            sub_cell.alignment = right_align
            sub_cell.number_format = '#,##0.00'
            sub_cell.fill = cat_fill

            subtotals[cat_name] = cat_total
            row += 1
            row += 1  # Spacing

        # ---- GRAND TOTALS ----
        grand_subtotal = sum(subtotals.values())
        tax_rate = 0.06
        tax_amount = round(grand_subtotal * tax_rate, 2)
        contingency_rate = 0.10
        contingency = round(grand_subtotal * contingency_rate, 2)
        grand_total = grand_subtotal + tax_amount + contingency

        totals_data = [
            ('SUBTOTAL', grand_subtotal),
            (f'TAX / GST ({int(tax_rate*100)}%)', tax_amount),
            (f'CONTINGENCY ({int(contingency_rate*100)}%)', contingency),
            ('GRAND TOTAL', grand_total),
        ]

        for label, val in totals_data:
            is_grand = label == 'GRAND TOTAL'
            ws.merge_cells(f'A{row}:G{row}')
            label_cell = ws.cell(row=row, column=1, value=label)
            label_cell.alignment = Alignment(horizontal='right', vertical='center')
            
            val_cell = ws.cell(row=row, column=8, value=val)
            val_cell.alignment = right_align
            val_cell.number_format = '#,##0.00'

            if is_grand:
                label_cell.font = total_font
                label_cell.fill = total_fill
                val_cell.font = total_font
                val_cell.fill = total_fill
                for c in range(1, 11):
                    ws.cell(row=row, column=c).fill = total_fill
            else:
                label_cell.font = Font(name="Segoe UI", size=10, bold=True, color="C9D1D9")
                label_cell.fill = cat_fill
                val_cell.font = money_font
                val_cell.fill = cat_fill

            row += 1

        # ---- FOOTER ----
        row += 1
        ws.merge_cells(f'A{row}:J{row}')
        ws.cell(row=row, column=1, value='This document is generated by Volts Design Platform.  All amounts are estimates subject to confirmation.')
        ws.cell(row=row, column=1).font = Font(name="Segoe UI", size=8, italic=True, color="484F58")

        # Freeze top rows
        ws.freeze_panes = 'A6'

        # Print setup
        ws.sheet_properties.pageSetUpPr = None
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1

        # Save
        filename = f"BoQ_{project_name}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        return filepath
